import os
import json
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from datetime import datetime
import google.generativeai as genai
from dotenv import load_dotenv
import re
from collections import defaultdict	

# 1. API 키 로드
load_dotenv()
genai.configure(api_key=os.getenv("GEMINI_API_KEY"))

# 2. 색상 설정
color_map = {
    "백엔드": "FF6666",
    "프론트엔드": "66B2FF",
    "하드웨어": "66CC99",
    "인공지능": "000000",
    "서류": "CCCCCC"
}

# 3. 프롬프트 생성 함수 (기획서 + 기능명세 포함된 JSON 기반)
def make_prompt(json_data: dict, parts: list, total_weeks: int, feedback: str = None) -> str:
    base_prompt = f"""
당신은 소프트웨어 기획 전문가이자 프로젝트 매니저입니다.

다음은 하나의 소프트웨어/서비스 프로젝트에 대한 기획서와 기능 명세서입니다. 당신의 임무는 아래 JSON을 바탕으로, 각 기능을 다음 개발 파트 중 하나 이상으로 분류하고, 프로젝트 기간 내에 모든 작업이 완료되도록 간트차트를 구성하는 것입니다.

✅ 개발 파트 목록:
- {', '.join(parts)}

✅ 당신이 수행할 작업:
1. 각 기능을 적절한 개발 파트에 할당하십시오 (중복 가능).
2. 논리적 선후 관계가 있다면 "선행작업"으로 표현하십시오.
3. 각 작업에 대해 개발 기간을 주 단위로 추정하십시오.
4. 전체 프로젝트 기간은 {total_weeks}주입니다. 그 안에 가능한 한 현실적인 일정으로 작업을 배분하십시오.
5. 기능 ID가 없다면 F-001, F-002... 형식으로 자동 생성하십시오.
6. 간트차트 출력 목적이므로 각 기능은 다음 항목을 포함해야 합니다:

[출력 JSON 형식 예시]
[
  {{
    "기획서요약": "해당 기능이 왜 필요한지 요약",
    "기능ID": "F-001",
    "기능명": "사용자 로그인",
    "파트": ["백엔드", "프론트엔드"],
    "기간": 2,
    "시작주차": 1,
    "선행작업": null
  }},
  ...
]
"""
    if feedback:
        base_prompt += f"\n\n📌 사용자 피드백 사항:\n{feedback}\n"

    base_prompt += f"\n✍️ 다음은 입력된 프로젝트 전체 JSON입니다:\n```json\n{json.dumps(json_data, ensure_ascii=False, indent=2)}\n```"
    return base_prompt

# 4. Gemini 호출 함수
def call_gemini(prompt):
    model = genai.GenerativeModel("gemini-2.5-flash")
    response = model.generate_content(prompt)
    return response.text.strip()

# 5. 사용자 입력
feature_file = input("📂 기능 명세서 JSON 파일명을 입력하세요 (예: features.json): ").strip()
if not os.path.exists(feature_file):
    print("❌ 파일이 존재하지 않습니다.")
    exit()

project_start = input("📅 프로젝트 시작일을 입력하세요 (예: 2025-08-12): ").strip()
try:
    datetime.strptime(project_start, "%Y-%m-%d")
except:
    print("❌ 날짜 형식이 올바르지 않습니다. (예: 2025-08-12)")
    exit()

total_weeks = input("📆 전체 프로젝트 기간(주 단위)을 입력하세요 (예: 12): ").strip()
if not total_weeks.isdigit():
    print("❌ 숫자를 입력해주세요.")
    exit()
total_weeks = int(total_weeks)

parts_input = input("🧩 사용할 개발 파트를 쉼표로 입력하세요 (예: 백엔드,프론트엔드,하드웨어,인공지능): ").strip()
available_parts = [p.strip() for p in parts_input.split(",")]
print(f"✅ 사용된 파트: {available_parts}")

# 6. JSON 로드 (기획서 + 기능명세 포함)
with open(feature_file, "r", encoding="utf-8") as f:
    json_data = json.load(f)

# Gemini 호출 및 피드백 루프
while True:
    prompt = make_prompt(json_data, available_parts, total_weeks)
    print("🚀 Gemini API 호출 중...")
    response_text = call_gemini(prompt)
    print("✅ 응답 완료")

    try:
        parsed_data = json.loads(response_text)
    except:
        json_str = re.search(r'\[\s*\{.*?\}\s*\]', response_text, re.DOTALL).group(0)
        parsed_data = json.loads(json_str)

    # 9. 간트차트 생성
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gantt Chart"

    weeks = [f"{i+1}주차" for i in range(total_weeks)]
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws["A1"] = "파트"
    ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
    ws["B1"] = "기능명"
    for i, w in enumerate(weeks):
        ws.cell(row=1, column=3+i).value = w
        ws.cell(row=1, column=3+i).alignment = Alignment(horizontal="center", vertical="center")

    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    current_row = 3

    grouped_tasks = defaultdict(list)
    for task in parsed_data:
        for part in task.get("파트", ["기타"]):
            grouped_tasks[part].append(task)

    for part in available_parts:
        if part not in grouped_tasks:
            continue
        tasks = grouped_tasks[part]
        part_start_row = current_row
        for task in tasks:
            name = task["기능명"]
            start = task.get("시작주차", 1)
            duration = task.get("기간", 1)

            ws.cell(row=current_row, column=1, value=part)
            ws.cell(row=current_row, column=2, value=name)

            for j in range(total_weeks):
                cell = ws.cell(row=current_row, column=3+j)
                if start-1 <= j < start-1 + duration:
                    cell.fill = PatternFill(start_color=color_map.get(part, "AAAAAA"), fill_type="solid")
                cell.border = border

            current_row += 1

        ws.merge_cells(start_row=part_start_row, start_column=1, end_row=current_row - 1, end_column=1)
        ws.cell(row=part_start_row, column=1).alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 30
    for col in range(3, 3+total_weeks):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 6

    # 저장 파일 이름 중복 방지
    base_filename = "간트차트_자동생성_업그레이드"
    ext = ".xlsx"
    filename = f"{base_filename}{ext}"
    counter = 1
    while os.path.exists(filename):
        filename = f"{base_filename}_{counter}{ext}"
        counter += 1

    wb.save(filename)
    print(f"🎉 간트차트 생성 완료 → {filename}")

    feedback = input("📣 간트차트 결과에 만족하시나요? (Y/N): ").strip().lower()
    if feedback == "n":
        comment = input("✏️ 재생성 시 반영할 피드백을 입력해주세요: ").strip()
        prompt = make_prompt(json_data, available_parts, total_weeks, feedback=comment)
        print("🔁 피드백을 반영하여 재생성 중...")
        continue
    else:
        break
