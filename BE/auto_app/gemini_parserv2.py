# -*- coding: utf-8 -*-
import os
import json
from dotenv import load_dotenv
import google.generativeai as genai
from google.generativeai.types import GenerationConfig
import datetime
import glob
from typing import Any, Dict, List

# 1. .env에서 API Key 로드
load_dotenv()
# API 키 환경 변수 이름을 "GOOGLE_API_KEY"로 통일합니다.
api_key = os.getenv("GEMINI_API_KEY_1")
if not api_key:
    raise ValueError("GEMINI_API_KEY_1 환경 변수가 설정되지 않았습니다.")
genai.configure(api_key=api_key)


def make_prompt(plan_text: str, existing_features: list = None) -> str:
    """
    Gemini 프롬프트를 생성합니다. 이미 추출된 기능 목록을 받아 중복을 방지합니다.
    """
    
    # 이전에 추출된 기능이 있는 경우, 프롬프트에 추가하여 중복을 방지하도록 명시
    if existing_features:
        deduplication_instruction = f"""
        ---
        🚨 **중요: 이전에 추출된 기능 목록**

        아래는 이미 추출된 기능 목록입니다. 기획서를 다시 한번 면밀히 검토하여, **아래 목록에 없는 새로운 기능들만 추가로 추출하십시오.**
        이미 있는 기능과 명칭이나 설명이 조금이라도 비슷하다면 절대 중복해서 생성해서는 안 됩니다.

        ```json
        {json.dumps(existing_features, indent=2, ensure_ascii=False)}
        ```
        ---
        """
        final_instruction = "이제 위 기획서를 다시 분석해, **이전에 추출되지 않은 새로운 기능 목록**을 최대한 많이 생성하십시오."
    else:
        deduplication_instruction = ""
        final_instruction = "이제 위 기획서를 분석해 위 JSON 구조에 따라 기능 목록을 최대한 많이 생성하십시오."

    return f'''
    다음은 한 개의 소프트웨어/서비스 기획서입니다.  
    당신은 이 문서를 분석하여, **기능 요구사항을 가능한 한 많이, 구체적으로, 그리고 정형화된 JSON 형식으로 추출**해야 합니다.

    🔍 **당신의 임무**

    1. 문서 전체를 끝까지 읽고, 해당 기획서에 포함된 기능을 모두 분리하여 나열하십시오.
    2. 하나의 기능 안에 여러 역할이 포함되어 있더라도, **논리적으로 분리 가능한 기능은 전부 독립된 기능으로 정의**하십시오.
    3. 기능은 아래 JSON 형식을 그대로 따르고, **가능한 모든 필드를 채우십시오.**

    {deduplication_instruction}

    🧱 **기능 JSON 출력 형식**

    ```json
    {{
      "기능ID": "FEAT-001",
      "기능명": "기능명을 간결하고 직관적으로 표현",
      "기능설명": {{
        "목적": "기능이 수행되는 이유와 필요성",
        "핵심역할": "실제로 수행하는 주된 작업 또는 처리"
      }},
      "사용자시나리오": {{
        "상황": "기능이 사용되는 사용자 상황",
        "행동": "사용자의 상호작용 방식"
      }},
      "입력값": {{ "필수": [], "선택": [], "형식": "" }},
      "출력값": {{ "요약정보": "", "상세정보": "" }},
      "처리방식": {{ "단계": [], "사용모델": "" }},
      "예외조건및처리": {{ "입력누락": "", "오류": "" }},
      "의존성또는연동항목": [],
      "기능우선순위": "높음 / 중간 / 낮음",
      "UI요소": [],
      "테스트케이스예시": []
    }}
    ```
    📌 주의사항:
    - **기능의 수에는 제한이 없습니다. 기획서에 명시적이거나 암시된 모든 기능을 최대한 많이 찾아서 포함하십시오.**
    - 오직 JSON 배열(리스트)로만 출력해야 하며, 다른 설명은 절대 추가하지 마십시오.

    ---
    📄 **기획서 원문**

    """
    {plan_text}
    """

    {final_instruction}
    **오직 JSON 배열로만 출력하십시오. 다른 설명은 출력하지 마십시오.**
    '''

def generate_feature_list(plan_text: str, existing_features: list = None):
    """
    Gemini를 호출하여 기능 목록을 생성합니다.
    """
    prompt = make_prompt(plan_text, existing_features)
    model = genai.GenerativeModel("gemini-1.5-flash") # 1.5-flash가 긴 컨텍스트 처리에 더 유리할 수 있음
    response = model.generate_content(prompt, generation_config=GenerationConfig(temperature=0.1))

    raw = response.text.strip()
    # 응답이 비어있는 경우 빈 리스트 반환
    if not raw:
        return []

    raw = raw.replace("```json", "").replace("```", "").strip()

    try:
        data = json.loads(raw)
        if isinstance(data, list):
            return data
        else:
            print("❌ 예상과 다른 구조(리스트가 아님):", type(data))
            return []
    except Exception as e:
        print("❌ JSON 파싱 실패:", e)
        print("🔎 원본 출력:\n", raw)
        return []

def _safe_get(mapping: Dict[str, Any], key: str, default: Any = "") -> Any:
    return mapping.get(key, default) if isinstance(mapping, dict) else default

def _to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    if isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, list):
        normalized: List[str] = []
        for item in value:
            if isinstance(item, (str, int, float)):
                normalized.append(str(item))
            else:
                try:
                    normalized.append(json.dumps(item, ensure_ascii=False))
                except Exception:
                    normalized.append(str(item))
        return ", ".join(normalized)
    if isinstance(value, dict):
        try:
            return json.dumps(value, ensure_ascii=False)
        except Exception:
            return str(value)
    return str(value)

def flatten_feature_to_row(feature: Dict[str, Any]) -> Dict[str, str]:
    """기능 JSON 한 건을 표 형태의 한 행으로 평탄화합니다."""
    desc = _safe_get(feature, "기능설명", {})
    scenario = _safe_get(feature, "사용자시나리오", {})
    inputs = _safe_get(feature, "입력값", {})
    outputs = _safe_get(feature, "출력값", {})
    process = _safe_get(feature, "처리방식", {})
    exceptions = _safe_get(feature, "예외조건및처리", {})

    row: Dict[str, str] = {
        "기능ID": _to_text(_safe_get(feature, "기능ID", "")),
        "기능명": _to_text(_safe_get(feature, "기능명", "")),
        "목적": _to_text(_safe_get(desc, "목적", "")),
        "핵심역할": _to_text(_safe_get(desc, "핵심역할", "")),
        "상황": _to_text(_safe_get(scenario, "상황", "")),
        "행동": _to_text(_safe_get(scenario, "행동", "")),
        "입력값_필수": _to_text(_safe_get(inputs, "필수", [])),
        "입력값_선택": _to_text(_safe_get(inputs, "선택", [])),
        "입력값_형식": _to_text(_safe_get(inputs, "형식", "")),
        "출력값_요약정보": _to_text(_safe_get(outputs, "요약정보", "")),
        "출력값_상세정보": _to_text(_safe_get(outputs, "상세정보", "")),
        "처리단계": _to_text(_safe_get(process, "단계", [])),
        "사용모델": _to_text(_safe_get(process, "사용모델", "")),
        "예외_입력누락": _to_text(_safe_get(exceptions, "입력누락", "")),
        "예외_오류": _to_text(_safe_get(exceptions, "오류", "")),
        "의존성": _to_text(_safe_get(feature, "의존성또는연동항목", [])),
        "기능우선순위": _to_text(_safe_get(feature, "기능우선순위", "")),
        "UI요소": _to_text(_safe_get(feature, "UI요소", [])),
        "테스트케이스예시": _to_text(_safe_get(feature, "테스트케이스예시", [])),
    }
    return row

def export_tabular_files(plan_lines: List[str], features: List[Dict[str, Any]], filename_base: str) -> None:
    """단일 엑셀 파일(xlsx)로 내보냅니다.

    - 시트 '기능목록': 전체 필드(정규 스키마)
    - 시트 '가독요약': 요약 표(가독성 중심)
    - 시트 '기획서원문': 원문 라인
    """
    try:
        import pandas as pd  # type: ignore
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        # 전체 필드 행 구성
        full_rows = [flatten_feature_to_row(f) for f in features]
        full_headers = [
            "기능ID",
            "기능명",
            "목적",
            "핵심역할",
            "상황",
            "행동",
            "입력값_필수",
            "입력값_선택",
            "입력값_형식",
            "출력값_요약정보",
            "출력값_상세정보",
            "처리단계",
            "사용모델",
            "예외_입력누락",
            "예외_오류",
            "의존성",
            "기능우선순위",
            "UI요소",
            "테스트케이스예시",
        ]

        # 요약 표 행 구성
        compact_rows = [_compact_row(f) for f in features]
        compact_headers = [
            "주요 기능",
            "페이지",
            "업무 대분류",
            "업무 중분류",
            "업무 소분류",
            "역할",
            "코멘트",
        ]

        xlsx_filename = f"{filename_base}.xlsx"
        with pd.ExcelWriter(xlsx_filename, engine="openpyxl") as writer:
            # 기능목록
            pd.DataFrame(full_rows, columns=full_headers).to_excel(
                writer, index=False, sheet_name="기능목록"
            )
            # 가독요약
            pd.DataFrame(compact_rows, columns=compact_headers).to_excel(
                writer, index=False, sheet_name="가독요약"
            )
            # 기획서원문
            pd.DataFrame({"기획서원문": plan_lines}).to_excel(
                writer, index=False, sheet_name="기획서원문"
            )

            # 서식 적용
            for sheet_name in ["기능목록", "가독요약"]:
                ws = writer.sheets[sheet_name]
                ws.freeze_panes = "A2"
                ws.auto_filter.ref = ws.dimensions
                # 헤더 스타일
                header_fill = PatternFill(start_color="FFE2E8F0", end_color="FFE2E8F0", fill_type="solid")
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.fill = header_fill
                    cell.alignment = Alignment(wrap_text=True, vertical="center")
                # 본문 정렬 및 자동 너비
                max_lengths: Dict[int, int] = {}
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for idx, cell in enumerate(row, start=1):
                        text = "" if cell.value is None else str(cell.value)
                        max_lengths[idx] = max(max_lengths.get(idx, 0), len(text))
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
                for col_idx, max_len in max_lengths.items():
                    col_letter = get_column_letter(col_idx)
                    ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 80)

        print(f"📊 엑셀 파일이 '{xlsx_filename}'로 저장되었습니다.")
    except ImportError:
        print("ℹ️ pandas/openpyxl이 없어 XLSX 저장을 건너뜁니다. 'pip install pandas openpyxl' 실행 후 다시 시도하세요.")
    except Exception as e:
        print("❌ XLSX 저장 실패:", e)

def _compact_row(feature: Dict[str, Any]) -> Dict[str, str]:
    """이미지 예시 형태(가독성 표)에 맞춘 요약 행 매핑."""
    desc = _safe_get(feature, "기능설명", {})
    scenario = _safe_get(feature, "사용자시나리오", {})

    # 페이지는 UI요소가 있으면 첫 항목, 없으면 '상황'을 대체 사용
    page_candidates = _safe_get(feature, "UI요소", [])
    page = ""
    if isinstance(page_candidates, list) and page_candidates:
        page = _to_text(page_candidates[0])
    elif _safe_get(scenario, "상황"):
        page = _to_text(_safe_get(scenario, "상황"))

    row = {
        "주요 기능": _to_text(_safe_get(feature, "기능명", "")),
        "페이지": page,
        "업무 대분류": _to_text(_safe_get(scenario, "상황", "")),
        "업무 중분류": _to_text(_safe_get(desc, "목적", "")),
        "업무 소분류": _to_text(_safe_get(scenario, "행동", "")),
        "역할": _to_text(_safe_get(desc, "핵심역할", "")),
        "코멘트": _to_text(_safe_get(_safe_get(feature, "출력값", {}), "요약정보", "")),
    }
    return row

# (요약 전용 개별 파일 내보내기 함수는 더 이상 사용하지 않습니다)

if __name__ == "__main__":
    print("✍️ 기획서를 입력하세요 (Enter 두 번으로 종료):")
    lines = []
    while True:
        line = input()
        if line.strip() == "":
            break
        lines.append(line)
    plan_text = "\n".join(lines)

    if not plan_text:
        print("입력된 내용이 없어 프로그램을 종료합니다.")
    else:
        final_features: List[Dict[str, Any]] = []
        MAX_PASSES = int(os.getenv("AUTO_PASSES", "5"))
        for pass_count in range(1, MAX_PASSES + 1):
            print("\n" + "="*50)
            print(f"🔍 자동 실행: 기능 추출 패스 #{pass_count} 진행")

            new_features = generate_feature_list(plan_text, existing_features=final_features)
            if new_features:
                start_id = len(final_features) + 1
                for i, feature in enumerate(new_features):
                    feature['기능ID'] = f"FEAT-{start_id+i:03d}"
                final_features.extend(new_features)
                print(f"✅ {len(new_features)}개 추가 → 누적 {len(final_features)}개")
            else:
                print("✅ 더 이상 새로운 기능을 찾지 못했습니다. 자동 종료합니다.")
                break

        # 최종 결과 저장
        print("\n" + "="*50)
        print(f"총 {len(final_features)}개의 기능으로 최종 명세서를 생성합니다.")
        
        existing_files = glob.glob("features_*.json")
        existing_nums = [int(f.split("_")[1].split(".")[0]) for f in existing_files if f.startswith("features_") and f.split("_")[1].split(".")[0].isdigit()]
        next_index = max(existing_nums) + 1 if existing_nums else 1
        filename = f"features_{next_index}.json"

        plan_lines = plan_text.splitlines()
        output_data = {
            "기획서원문": plan_lines,
            "기능목록": final_features
        }

        try:
            with open(filename, "w", encoding="utf-8") as f:
                json.dump(output_data, f, ensure_ascii=False, indent=2)
            print(f"\n💾 최종 기능 명세서가 '{filename}' 파일에 저장되었습니다.")
        except Exception as e:
            print("❌ JSON 저장 실패:", e)
        
        # CSV / XLSX로도 내보내기
        base = filename.rsplit(".", 1)[0]
        export_tabular_files(plan_lines, final_features, base)
