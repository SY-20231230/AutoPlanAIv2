import os
import json
from typing import Dict, List, Optional
from dotenv import load_dotenv
import google.generativeai as genai
from google.generativeai.types import GenerationConfig

# 1. 환경 변수 로드 및 Gemini 설정
load_dotenv()
# GOOGLE_API_KEY 또는 GEMINI_API_KEY_2를 사용하도록 유연하게 변경
api_key = os.getenv("GEMINI_API_KEY_2") or os.getenv("GOOGLE_API_KEY")
if api_key:
    genai.configure(api_key=api_key)

# ======================================================================
# ✅ 2. Gemini 프롬프트 생성 함수 (수정된 부분)
# ======================================================================
def make_refine_prompt(plan_text, feature_list):
    """
    Gemini 2 정제용 프롬프트를 생성합니다.
    AI가 원래의 JSON 구조를 반드시 유지하도록 명확하게 지시합니다.
    """
    return f"""
당신은 뛰어난 시스템 분석가입니다. 주어진 기획서 원문과 1차로 추출된 기능 명세 목록을 검토하여, 각 기능의 내용을 더 명확하고 전문적으로 개선하는 임무를 맡았습니다.

---
📄 **기획서 원문 (참고용)**:
{plan_text}
---
🧱 **1차 추출된 기능 명세 목록 (수정 대상)**:
```json
{json.dumps(feature_list, indent=2, ensure_ascii=False)}
```
---

### 🚨 **당신의 핵심 임무**

1.  **구조 절대 유지:** 위에 주어진 **"1차 추출된 기능 명세 목록"의 JSON 구조와 모든 키(key) 이름("기능ID", "기능명" 등)을 절대로 변경하거나 삭제해서는 안 됩니다.**
2.  **내용 개선:** 각 키에 해당하는 값(value)의 내용을 기획서 원문을 참고하여 더 구체적이고, 일관성 있으며, 전문적인 용어로 다듬어주세요. 비어있는 필드가 있다면 최대한 채워주세요.
3.  **형식 준수:** 최종 결과물은 **원래와 동일한 구조의 JSON 배열(리스트) 형식**으로만 출력해야 합니다. 다른 설명이나 Markdown 코드 블록(` ```json`)은 절대 추가하지 마십시오.

이제 위 규칙을 반드시 준수하여, 아래 기능 명세 목록을 정제한 최종 JSON 배열을 출력하십시오.
"""

# 3. 메인 실행 함수 (테스트용)
def refine_features():
    # ... (이하 코드는 기존과 동일하게 유지) ...
    input_file = input("📂 불러올 JSON 파일명을 입력하세요 (예: features_2.json): ").strip()

    try:
        with open(input_file, "r", encoding="utf-8") as f:
            data = json.load(f)
            plan_text = "\n".join(data["기획서원문"])
            feature_list = data["기능목록"]
    except Exception as e:
        print("❌ 입력 파일 로딩 실패:", e)
        return

    # Gemini 호출
    prompt = make_refine_prompt(plan_text, feature_list)
    model = genai.GenerativeModel("gemini-1.5-flash") # 모델명은 상황에 맞게 조정 가능
    response = model.generate_content(prompt, generation_config=GenerationConfig(temperature=0.2, response_mime_type="application/json"))
    result_text = response.text.strip()

    try:
        refined_json = json.loads(result_text)
    except json.JSONDecodeError as e:
        print(f"❌ Gemini 응답을 JSON으로 파싱하는 데 실패했습니다: {e}")
        print("--- Gemini 원본 응답 ---")
        print(result_text)
        print("--------------------")
        print("⚠️ 원본 텍스트를 그대로 파일에 저장합니다.")
        refined_json = result_text # 파싱 실패 시 원본 텍스트 사용


    # 저장할 파일 이름: 입력 파일명 + _fix.json
    base_input_name = os.path.splitext(os.path.basename(input_file))[0]
    filename = f"{base_input_name}_fix.json"

    # 결과 저장
    try:
        with open(filename, "w", encoding="utf-8") as f:
            # G2는 이제 리스트를 반환하므로, 래핑 없이 바로 저장하거나 필요에 따라 래핑
            json.dump(refined_json, f, ensure_ascii=False, indent=2)
        print(f"✅ 정제된 기획서가 '{filename}'에 저장되었습니다.")
    except Exception as e:
        print("❌ 파일 저장 실패:", e)

    # 엑셀 동기화 저장
    try:
        plan_lines = data.get("기획서원문", [])
        # AI가 구조를 유지했으므로 refined_json을 바로 사용
        features_for_excel = refined_json if isinstance(refined_json, list) else []
        if features_for_excel:
            base_input_name = os.path.splitext(os.path.basename(input_file))[0]
            export_excel_from_features(plan_lines, features_for_excel, f"{base_input_name}_fix.xlsx")
        else:
            print("ℹ️ 정제된 기능 목록이 없어 엑셀 동기화를 건너뜁니다.")
    except Exception as e:
        print("❌ 엑셀 동기화 저장 중 오류:", e)


# --------------------------- Excel 동기화 유틸 ---------------------------
def _safe_get(mapping, key, default=""):
    return mapping.get(key, default) if isinstance(mapping, dict) else default

def _to_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    if isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, list):
        items = []
        for v in value:
            if isinstance(v, (str, int, float)):
                items.append(str(v))
            else:
                try:
                    items.append(json.dumps(v, ensure_ascii=False))
                except Exception:
                    items.append(str(v))
        return ", ".join(items)
    if isinstance(value, dict):
        try:
            return json.dumps(value, ensure_ascii=False)
        except Exception:
            return str(value)
    return str(value)

def flatten_feature_to_row(feature: dict) -> dict:
    desc = _safe_get(feature, "기능설명", {})
    scenario = _safe_get(feature, "사용자시나리오", {})
    inputs = _safe_get(feature, "입력값", {})
    outputs = _safe_get(feature, "출력값", {})
    process = _safe_get(feature, "처리방식", {})
    exceptions = _safe_get(feature, "예외조건및처리", {})

    return {
        "기능ID": _to_text(_safe_get(feature, "기능ID", "")),
        "기능명": _to_text(_safe_get(feature, "기능명", "")),
        "목적": _to_text(_safe_get(desc, "목적", "")),
        "핵심역할": _to_text(_safe_get(desc, "핵심역할", "")),
        "상황": _to_text(_safe_get(scenario, "상황", "")),
        "행동": _to_text(_safe_get(scenario, "행동", "")),
        "입력값_필수": _to_text(_safe_get(inputs, "필수", [])),
        "입력값_선택": _to_text(_safe_get(inputs, "선택", [])),
        "입력값_형식": _to_text(_safe_get(inputs, "형식", "")),
        "출력값_요약정보": _to_text(_safe_get(outputs, "요약정보", "")),
        "출력값_상세정보": _to_text(_safe_get(outputs, "상세정보", "")),
        "처리단계": _to_text(_safe_get(process, "단계", [])),
        "사용모델": _to_text(_safe_get(process, "사용모델", "")),
        "예외_입력누락": _to_text(_safe_get(exceptions, "입력누락", "")),
        "예외_오류": _to_text(_safe_get(exceptions, "오류", "")),
        "의존성": _to_text(_safe_get(feature, "의존성또는연동항목", [])),
        "기능우선순위": _to_text(_safe_get(feature, "기능우선순위", "")),
        "UI요소": _to_text(_safe_get(feature, "UI요소", [])),
        "테스트케이스예시": _to_text(_safe_get(feature, "테스트케이스예시", [])),
    }

def _compact_row(feature: dict) -> dict:
    desc = _safe_get(feature, "기능설명", {})
    scenario = _safe_get(feature, "사용자시나리오", {})
    page_candidates = _safe_get(feature, "UI요소", [])
    page = ""
    if isinstance(page_candidates, list) and page_candidates:
        page = _to_text(page_candidates[0])
    elif _safe_get(scenario, "상황"):
        page = _to_text(_safe_get(scenario, "상황"))
    return {
        "주요 기능": _to_text(_safe_get(feature, "기능명", "")),
        "페이지": page,
        "업무 대분류": _to_text(_safe_get(scenario, "상황", "")),
        "업무 중분류": _to_text(_safe_get(desc, "목적", "")),
        "업무 소분류": _to_text(_safe_get(scenario, "행동", "")),
        "역할": _to_text(_safe_get(desc, "핵심역할", "")),
        "코멘트": _to_text(_safe_get(_safe_get(feature, "출력값", {}), "요약정보", "")),
    }

def export_excel_from_features(plan_lines: List[str], features: List[dict], out_path: str) -> None:
    try:
        import pandas as pd
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("ℹ️ pandas/openpyxl이 없어 XLSX 생성을 건너뜁니다. 'pip install pandas openpyxl' 후 재시도하세요.")
        return

    full_headers = [
        "기능ID","기능명","목적","핵심역할","상황","행동",
        "입력값_필수","입력값_선택","입력값_형식","출력값_요약정보","출력값_상세정보",
        "처리단계","사용모델","예외_입력누락","예외_오류","의존성","기능우선순위","UI요소","테스트케이스예시",
    ]
    compact_headers = ["주요 기능","페이지","업무 대분류","업무 중분류","업무 소분류","역할","코멘트"]

    full_rows = [flatten_feature_to_row(f) for f in features]
    compact_rows = [_compact_row(f) for f in features]

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        pd.DataFrame(full_rows, columns=full_headers).to_excel(writer, index=False, sheet_name="기능목록")
        pd.DataFrame(compact_rows, columns=compact_headers).to_excel(writer, index=False, sheet_name="가독요약")
        pd.DataFrame({"기획서원문": plan_lines}).to_excel(writer, index=False, sheet_name="기획서원문")

        for sheet_name in ["기능목록", "가독요약"]:
            ws = writer.sheets[sheet_name]
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(wrap_text=True, vertical="center")
            max_lengths = {}
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for idx, cell in enumerate(row, start=1):
                    text = "" if cell.value is None else str(cell.value)
                    max_lengths[idx] = max(max_lengths.get(idx, 0), len(text))
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
            for col_idx, max_len in max_lengths.items():
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 80)

    print(f"📊 엑셀 동기화 완료 → '{out_path}'")

# 실행
if __name__ == "__main__":
    refine_features()
