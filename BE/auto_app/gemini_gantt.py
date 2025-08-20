# gemini_gantt.py
# DB의 확정된 Requirement를 기반으로 간트차트를 생성하기 위한 유틸 함수 모음
# - 외부에서(views.py) project 객체와 Requirement queryset을 받아 prompt 생성
# - Gemini 호출 후 응답(JSON 배열) 파싱
# - openpyxl로 .xlsx 생성

import os
import json
import re
from datetime import datetime
from collections import defaultdict

import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

import google.generativeai as genai
from dotenv import load_dotenv

# ─────────────────────────────────────────────────────────────
# 1) 환경변수/LLM 설정
# ─────────────────────────────────────────────────────────────
load_dotenv()  # .env 에서 GEMINI_API_KEY_3 로드
genai.configure(api_key=os.getenv("GEMINI_API_KEY_3"))

# ─────────────────────────────────────────────────────────────
# 2) 색상 맵 (파트별 채우기 색상)
# ─────────────────────────────────────────────────────────────
COLOR_MAP = {
    "백엔드": "FF6666",
    "프론트엔드": "66B2FF",
    "하드웨어": "66CC99",
    "인공지능": "9999FF",
    "서류": "CCCCCC",
}

# ─────────────────────────────────────────────────────────────
# 3) DB → 프롬프트 입력용 JSON 변환
#    - 프로젝트/요구사항 모델을 직접 import하지 않고, duck-typing으로 접근
#    - views에서 Requirement.objects.filter(...confirmed_by_user=True) 전달
# ─────────────────────────────────────────────────────────────
def build_payload_from_db(project, requirements_queryset):
    """
    project: Project 인스턴스 (title, description 속성 사용)
    requirements_queryset: Requirement 쿼리셋
      - 각 객체는 다음 속성을 가정:
        .Requirement (PK), .feature_name, .summary, .description(JSON 문자열일 수도), .project

    반환값 예시:
    {
      "project": {"title": "...", "description": "..."},
      "features": [
        {"기능ID": 1, "기능명": "로그인", "요약": "...", "원본": {...}},
        ...
      ]
    }
    """
    items = []
    for r in requirements_queryset:
        # description에 JSON이 있을 수 있으므로 안전 파싱
        src = {}
        if getattr(r, "description", None):
            try:
                src = json.loads(r.description)
            except Exception:
                src = {}
        items.append({
            "기능ID": getattr(r, "Requirement", None),
            "기능명": getattr(r, "feature_name", "") or "기능",
            "요약": getattr(r, "summary", "") or "",
            "원본": src,
        })

    payload = {
        "project": {
            "title": getattr(project, "title", "") or "",
            "description": getattr(project, "description", "") or "",
        },
        "features": items,
    }
    return payload

# ─────────────────────────────────────────────────────────────
# 4) 프롬프트 생성
# ─────────────────────────────────────────────────────────────
def make_prompt(json_data: dict, parts: list, total_weeks: int, feedback: str = None) -> str:
    """
    json_data: build_payload_from_db 로 생성된 dict
    parts: ["백엔드","프론트엔드", ...]
    total_weeks: 전체 프로젝트 기간(주)
    feedback: 재생성 시 반영할 코멘트(선택)
    """
    base_prompt = f"""
당신은 소프트웨어 기획 전문가이자 프로젝트 매니저입니다.

아래는 하나의 프로젝트에 대한 기획서와 확정된 기능명세서 목록입니다.
당신의 임무는 이 정보를 바탕으로, 지정된 개발 파트들 중 하나 이상에 각 기능을 배정하고,
전체 프로젝트 기간 {total_weeks}주 안에서 가능한 현실적인 간트차트를 구성하는 것입니다.

✅ 개발 파트 목록:
- {", ".join(parts)}

✅ 지켜야 할 규칙:
1) 각 기능을 적절한 개발 파트(복수 가능)에 배정하세요.
2) 선후 관계가 있다면 "선행작업" 필드에 기능ID 목록 또는 단일 값으로 표현하세요. (없으면 null)
3) 각 기능의 기간은 "주" 단위 정수로 추정하세요.
4) "시작주차"는 1부터 시작하는 정수입니다.
5) 기능ID가 없다면 F-001, F-002... 형식으로 자동 생성하세요.
6) 출력은 반드시 JSON 배열만 반환하세요. (설명 텍스트 금지)

[출력 JSON 형식 예시]
[
  {{
    "기획서요약": "해당 기능이 왜 필요한지 한 줄 요약",
    "기능ID": "F-001",
    "기능명": "사용자 로그인",
    "파트": ["백엔드","프론트엔드"],
    "기간": 2,
    "시작주차": 1,
    "선행작업": null
  }},
  ...
]
""".strip()

    if feedback:
        base_prompt += f"\n\n📌 사용자 피드백:\n{feedback}"

    # 입력 데이터를 JSON으로 포함
    base_prompt += f"\n\n✍️ 프로젝트/기능 전체 데이터(JSON):\n```json\n{json.dumps(json_data, ensure_ascii=False, indent=2)}\n```"
    return base_prompt

# ─────────────────────────────────────────────────────────────
# 5) Gemini 호출
# ─────────────────────────────────────────────────────────────
def call_gemini(prompt: str) -> str:
    """
    Gemini 2.5 Flash 호출. text 결과만 반환.
    """
    model = genai.GenerativeModel("gemini-1.5-flash")
    resp = model.generate_content(prompt)
    return (resp.text or "").strip()

# ─────────────────────────────────────────────────────────────
# 6) LLM 응답(JSON 배열) 파싱
# ─────────────────────────────────────────────────────────────
def parse_llm_array(text: str):
    """
    LLM이 JSON 외 설명을 섞어줄 수 있으므로, 배열 부분만 안전 추출.
    반환: list (작업 항목들)
    """
    try:
        data = json.loads(text)
        if isinstance(data, list):
            return data
        # dict 등으로 감싸져 온 경우 배열 후보 추출
    except Exception:
        pass

    m = re.search(r"\[\s*\{.*?\}\s*\]", text, re.DOTALL)
    if not m:
        raise ValueError("LLM 응답에서 유효한 JSON 배열을 찾지 못했습니다.")
    return json.loads(m.group(0))

# ─────────────────────────────────────────────────────────────
# 7) 간트차트 엑셀(.xlsx) 생성
# ─────────────────────────────────────────────────────────────
def build_gantt_xlsx(parsed_data: list, total_weeks: int, available_parts: list, save_path: str):
    """
    parsed_data 예시 원소:
    {
      "기획서요약": "...",
      "기능ID": "F-001" 또는 실제 ID,
      "기능명": "로그인",
      "파트": ["백엔드","프론트엔드"],
      "기간": 2,
      "시작주차": 1,
      "선행작업": null 또는 ["F-000"] 등
    }
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gantt Chart"

    # 헤더(2행 구조)
    weeks = [f"{i+1}주차" for i in range(int(total_weeks))]
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws["A1"] = "파트"
    ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)
    ws["B1"] = "기능명"

    for i, w in enumerate(weeks):
        c = ws.cell(row=1, column=3+i, value=w)
        c.alignment = Alignment(horizontal="center", vertical="center")

    # 셀 테두리
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # 파트별 그룹핑
    grouped = defaultdict(list)
    for task in parsed_data:
        parts = task.get("파트") or ["기타"]
        for p in parts:
            grouped[str(p)].append(task)

    # 본문 그리기
    row = 3
    for part in available_parts:
        part = str(part)
        tasks = grouped.get(part, [])
        if not tasks:
            continue

        start_row = row
        for t in tasks:
            name = t.get("기능명") or t.get("기능ID") or "작업"
            start = int(t.get("시작주차", 1))
            duration = int(t.get("기간", 1))

            ws.cell(row=row, column=1, value=part)
            ws.cell(row=row, column=2, value=name)

            for j in range(int(total_weeks)):
                cell = ws.cell(row=row, column=3+j)
                # 간트 바 채우기
                if start-1 <= j < start-1 + duration:
                    cell.fill = PatternFill(
                        start_color=COLOR_MAP.get(part, "AAAAAA"),
                        fill_type="solid"
                    )
                cell.border = border
            row += 1

        # 파트 머지/정렬
        ws.merge_cells(start_row=start_row, start_column=1, end_row=row-1, end_column=1)
        ws.cell(row=start_row, column=1).alignment = Alignment(horizontal="center", vertical="center")

    # 컬럼 너비
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 30
    for col in range(3, 3 + int(total_weeks)):
        ws.column_dimensions[get_column_letter(col)].width = 6

    # 저장
    wb.save(save_path)

# ─────────────────────────────────────────────────────────────
# 8) (선택) 유틸: 고유 파일명 만들기
# ─────────────────────────────────────────────────────────────
def unique_filename(outdir: str, base: str = "간트차트", ext: str = ".xlsx") -> str:
    """
    같은 폴더에 같은 파일명이 있으면 뒤에 _1, _2... 붙여서 충돌 방지
    """
    name = f"{base}{ext}"
    i = 1
    while os.path.exists(os.path.join(outdir, name)):
        name = f"{base}_{i}{ext}"
        i += 1
    return name

# ─────────────────────────────────────────────────────────────
# (모듈로만 사용) - CLI/직접 실행 방지
# ─────────────────────────────────────────────────────────────
if __name__ == "__main__":
    # 장고에서 import 하여 사용하도록 설계. 직접 실행 로직 없음.
    pass
